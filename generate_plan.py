import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "New Tasks"

# Column headers matching user's project plan
headers = ["WBS", "Tasks/Activities", "Responsible", "Start", "Finish", "", "% Tamamlanma", "Work Days"]

# Week headers (Week 9 to Week 16) - 5 days each
day_labels = ["P", "S", "C", "P", "C"]  # Mon-Fri Turkish
for w in range(9, 17):
    for d in day_labels:
        headers.append(d)

# Tasks scheduled March 30 - May 20, 2026
# Format: (WBS, Task, Responsible, Start, Finish, Duration, %, WorkDays, week_day_pairs)
tasks = [
    ("2.0", "Testing & Quality Assurance", "Mahmut & Eren", "30.03.2026", "13.04.2026", 11, "0%", 11, []),
    ("2.1", "Unit Tests", "Mahmut Muhammed Pakim", "30.03.2026", "02.04.2026", 4, "0%", 4,
     [(9,1),(9,2),(9,3),(9,4)]),
    ("2.2", "Integration Tests", "Mahmut & Eren", "03.04.2026", "08.04.2026", 4, "0%", 4,
     [(9,5),(10,1),(10,2),(10,3)]),
    ("2.3", "User Acceptance Testing (UAT)", "Eren Görkem Çolak", "09.04.2026", "13.04.2026", 3, "0%", 3,
     [(10,4),(10,5),(11,1)]),
    ("3.0", "Improvements & Updates", "Mahmut & Eren", "14.04.2026", "01.05.2026", 14, "0%", 14, []),
    ("3.1", "Security Improvements", "Mahmut Muhammed Pakim", "14.04.2026", "16.04.2026", 3, "0%", 3,
     [(11,2),(11,3),(11,4)]),
    ("3.2", "Performance Optimization", "Eren Görkem Çolak", "17.04.2026", "21.04.2026", 3, "0%", 3,
     [(11,5),(12,1),(12,2)]),
    ("3.3", "Bug Fixes", "Mahmut & Eren", "22.04.2026", "24.04.2026", 3, "0%", 3,
     [(12,3),(12,4),(12,5)]),
    ("3.4", "UI/UX Updates", "Eren Görkem Çolak", "27.04.2026", "29.04.2026", 3, "0%", 3,
     [(13,1),(13,2),(13,3)]),
    ("3.5", "New Feature Additions", "Mahmut Muhammed Pakim", "30.04.2026", "01.05.2026", 2, "0%", 2,
     [(13,4),(13,5)]),
    ("4.0", "Documentation", "Mahmut & Eren", "04.05.2026", "13.05.2026", 8, "0%", 8, []),
    ("4.1", "Technical Documentation", "Mahmut Muhammed Pakim", "04.05.2026", "06.05.2026", 3, "0%", 3,
     [(14,1),(14,2),(14,3)]),
    ("4.2", "User Guide", "Eren Görkem Çolak", "07.05.2026", "08.05.2026", 2, "0%", 2,
     [(14,4),(14,5)]),
    ("4.3", "Project Report", "Mahmut & Eren", "11.05.2026", "13.05.2026", 3, "0%", 3,
     [(15,1),(15,2),(15,3)]),
    ("5.0", "Final Delivery & Presentation", "Mahmut & Eren", "14.05.2026", "20.05.2026", 5, "0%", 5, []),
    ("5.1", "Final Presentation Preparation", "Mahmut & Eren", "14.05.2026", "18.05.2026", 3, "0%", 3,
     [(15,4),(15,5),(16,1)]),
    ("5.2", "Project Submission", "Mahmut & Eren", "19.05.2026", "20.05.2026", 2, "0%", 2,
     [(16,2),(16,3)]),
]

# Styles
thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF", size=10)
parent_fill = PatternFill("solid", fgColor="D9E2F3")
parent_font = Font(bold=True, size=10)
gantt_fill = PatternFill("solid", fgColor="70AD47")  # Green for active days

# Write headers
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add week number headers (row above day letters for clarity)
# Week 9 starts at col 9, each week = 5 cols
ws.insert_rows(1)
for w_idx, w_num in enumerate(range(9, 17)):
    start_col = 9 + w_idx * 5
    cell = ws.cell(row=1, column=start_col, value=f"Week {w_num}")
    cell.font = Font(bold=True, size=9)
    cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+4)

# Write data (starting row 3)
for row_idx, task in enumerate(tasks, 3):
    wbs, name, resp, start, finish, dur, pct, wdays, gantt_days = task
    
    values = [wbs, name, resp, start, finish, dur, pct, wdays]
    is_parent = wbs.endswith(".0")
    
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        # WBS code as text
        if col == 1:
            cell.number_format = '@'
        if is_parent:
            cell.fill = parent_fill
            cell.font = parent_font
    
    # Fill empty day columns with border
    for col in range(9, 9 + 40):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
    
    # Mark active days with green fill
    for week, day in gantt_days:
        col = 9 + (week - 9) * 5 + (day - 1)
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = gantt_fill
        cell.border = thin_border

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 26
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 5
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 10
# Day columns narrow
from openpyxl.utils import get_column_letter
for i in range(9, 49):
    ws.column_dimensions[get_column_letter(i)].width = 3

ws.freeze_panes = 'A3'

output = "/Users/muhammedpakim/Documents/e-commerce/Eventix_ProjectPlan_NewTasks.xlsx"
wb.save(output)
print(f"Dosya oluşturuldu: {output}")
