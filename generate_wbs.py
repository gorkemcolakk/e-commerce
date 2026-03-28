import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "WBS"

headers = ["WBS Code", "Level", "Deliverable / Work Package Name", "Description",
           "Responsible Person / Team", "Acceptance Criteria",
           "Estimated Duration (days)", "Estimated Cost ($)", "Dependencies", "Notes"]

data = [
    ["1.0", 1, "Eventix Ticketing Platform", "Complete event e-commerce system", "Mahmut & Eren", "All deliverables accepted", 30, 1500, "-", "Student Project"],
    ["1.1", 2, "User Management System", "Registration, login, user profiles", "Mahmut & Eren", "Auth system works smoothly", 5, 200, "-", "Student Project"],
    ["1.1.1", 3, "Registration & Login Backend", "JWT auth, registration, validation", "Mahmut Muhammed Pakim", "Users can securely log in", 3, 100, "-", "Student Project"],
    ["1.1.2", 3, "Profile & Roles UI", "Frontend forms, profile updates", "Eren Görkem Çolak", "Profiles update successfully", 2, 100, "1.1.1", "Student Project"],
    ["1.2", 2, "Event Catalog System", "Event listing, categories, structure", "Mahmut & Eren", "Events load without errors", 7, 350, "-", "Student Project"],
    ["1.2.1", 3, "Event CRUD & Database", "SQLite DB schema, CRUD endpoints", "Mahmut Muhammed Pakim", "Backend API runs properly", 3, 150, "-", "Student Project"],
    ["1.2.2", 3, "Frontend Listing & Details Page", "Event cards, search, event details UI", "Eren Görkem Çolak", "Responsive UI displays events", 4, 200, "1.2.1", "Student Project"],
    ["1.3", 2, "Payment & Ticketing Flow", "Checkout, virtual POS, seating", "Mahmut & Eren", "Tickets can be purchased", 6, 400, "1.2", "Student Project"],
    ["1.3.1", 3, "Payment & Checkout Logic", "Virtual POS simulation, promo codes", "Mahmut Muhammed Pakim", "Payments process accurately", 3, 200, "1.1.1", "Student Project"],
    ["1.3.2", 3, "QR Generation & Seat Selection", "Seat plan UI, QR code ticket creation", "Eren Görkem Çolak", "Seat map works, QR generated", 3, 200, "1.2.1", "Student Project"],
    ["1.4", 2, "Reporting & Dashboards", "Admin and Organizer panels", "Mahmut & Eren", "Management panels load", 7, 350, "1.2", "Student Project"],
    ["1.4.1", 3, "Organizer Rev & Analytics", "Gross/Net revenue calculation logic", "Mahmut Muhammed Pakim", "Correct commission/revenue output", 4, 200, "1.3", "Student Project"],
    ["1.4.2", 3, "Admin Event Approval UI", "Admin user/event lists, approve/reject", "Eren Görkem Çolak", "Admin can approve/reject events", 3, 150, "1.2.1", "Student Project"],
    ["1.5", 2, "Platform Operations", "Emails, notifications, and UI finish", "Mahmut & Eren", "System alerts trigger correctly", 5, 200, "-", "Student Project"],
    ["1.5.1", 3, "Email Service & Notifications", "SMTP mailer, password resets", "Mahmut Muhammed Pakim", "Emails are sent on triggers", 3, 100, "-", "Student Project"],
    ["1.5.2", 3, "UI/UX Polish & Mobile CSS", "Overall glassmorphism, responsive UI", "Eren Görkem Çolak", "No visual bugs on mobile", 2, 100, "-", "Student Project"],
    ["2.0", 1, "Test & Kalite Güvence", "Tüm sistemin test edilmesi ve hata giderme", "Mahmut & Eren", "Tüm testler başarılı", 10, 400, "1.0", "Student Project"],
    ["2.1", 2, "Birim Testleri (Unit Tests)", "Backend API endpoint'leri için birim testleri", "Mahmut Muhammed Pakim", "Tüm endpoint'ler test edilmiş", 4, 150, "1.0", "Student Project"],
    ["2.2", 2, "Entegrasyon Testleri", "Modüller arası entegrasyon testleri (ödeme-bilet-email akışı)", "Mahmut & Eren", "End-to-end akışlar sorunsuz çalışıyor", 3, 150, "2.1", "Student Project"],
    ["2.3", 2, "Kullanıcı Kabul Testleri (UAT)", "Gerçek kullanıcılar ile test senaryolarının yürütülmesi", "Eren Görkem Çolak", "Kullanıcı geri bildirimleri olumlu", 3, 100, "2.2", "Student Project"],
    ["3.0", 1, "İyileştirme & Güncelleme", "Güvenlik, performans ve yeni özellik güncellemeleri", "Mahmut & Eren", "Tüm iyileştirmeler uygulanmış", 12, 500, "1.0", "Student Project"],
    ["3.1", 2, "Güvenlik İyileştirmeleri", "SQL injection, XSS, CSRF korumaları, input validasyon", "Mahmut Muhammed Pakim", "Bilinen güvenlik açıkları kapatılmış", 3, 100, "1.0", "Student Project"],
    ["3.2", 2, "Performans Optimizasyonu", "Veritabanı sorgu optimizasyonu, caching, sayfa yükleme hızı", "Eren Görkem Çolak", "Sayfa yükleme süresi < 3 saniye", 2, 100, "1.0", "Student Project"],
    ["3.3", 2, "Hata Düzeltmeleri (Bug Fixes)", "Test aşamasında tespit edilen hataların giderilmesi", "Mahmut & Eren", "Bilinen tüm buglar çözülmüş", 2, 100, "2.0", "Student Project"],
    ["3.4", 2, "Kullanıcı Arayüzü Güncellemeleri", "Kullanıcı geri bildirimlerine göre UI/UX iyileştirmeleri", "Eren Görkem Çolak", "Güncellenmiş UI kullanıcı onayı almış", 3, 100, "2.3", "Student Project"],
    ["3.5", 2, "Yeni Özellik Eklemeleri", "Wishlist, gelişmiş filtreleme, bildirim tercihleri", "Mahmut Muhammed Pakim", "Yeni özellikler sorunsuz çalışıyor", 2, 100, "1.0", "Student Project"],
    ["4.0", 1, "Dokümantasyon", "Proje teknik ve kullanıcı dokümantasyonu", "Mahmut & Eren", "Dokümantasyon eksiksiz teslim edilmiş", 7, 250, "1.0", "Student Project"],
    ["4.1", 2, "Teknik Dokümantasyon", "API dokümantasyonu, veritabanı şeması, mimari diyagramlar", "Mahmut Muhammed Pakim", "API endpointleri ve DB şeması belgelenmiş", 3, 100, "1.0", "Student Project"],
    ["4.2", 2, "Kullanıcı Kılavuzu", "Son kullanıcı için platform kullanım rehberi", "Eren Görkem Çolak", "Tüm kullanıcı rolleri için kılavuz hazır", 2, 75, "1.0", "Student Project"],
    ["4.3", 2, "Proje Raporu", "Final proje raporu, metodoloji ve bulgular", "Mahmut & Eren", "Rapor akademik standartlara uygun", 2, 75, "4.1, 4.2", "Student Project"],
    ["5.0", 1, "Final Teslim & Sunum", "Projenin son teslimi ve sunum hazırlığı", "Mahmut & Eren", "Proje başarıyla teslim edilmiş", 5, 100, "2.0, 3.0, 4.0", "Student Project"],
    ["5.1", 2, "Final Sunum Hazırlığı", "Demo senaryosu, sunum slaytları, canlı demo hazırlığı", "Mahmut & Eren", "Sunum materyalleri hazır", 3, 50, "5.0", "Student Project"],
    ["5.2", 2, "Proje Teslimi", "Kaynak kod, dokümantasyon ve raporun final teslimi", "Mahmut & Eren", "Tüm çıktılar teslim edilmiş", 2, 50, "5.1", "Student Project"],
]

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
level1_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
level1_font = Font(bold=True, size=11)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Write data - WBS Code column as TEXT
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # WBS Code (col 1) and Dependencies (col 9) as text
        if col_idx == 1 or col_idx == 9:
            cell.number_format = '@'  # Force text format
        
        # Level 1 rows bold with background
        level = row_data[1]
        if level == 1:
            cell.fill = level1_fill
            cell.font = level1_font

# Column widths
col_widths = [12, 6, 35, 50, 28, 38, 22, 18, 22, 14]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze top row
ws.freeze_panes = 'A2'

output_path = "/Users/muhammedpakim/Documents/e-commerce/Eventix_WBS.xlsx"
wb.save(output_path)
print(f"WBS Excel dosyası oluşturuldu: {output_path}")
